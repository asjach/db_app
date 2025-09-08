from PySide6.QtWidgets import QDialog, QMessageBox
from ui.ui_dialog_input_excel import Ui_Form
from models.model_preferensi import Model_Preferensi
# from utils.static_values import LEFT_COLUMN, KOLOM_ANGKA, KOLOM_CURRENCY, KOLOM_FLOAT, KOLOM_TANGGAL
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from utils.fungsi.general_functions import *
import os
import pandas as pd
import numpy as np
# from PySide6.QtWidgets import QApplication

class DialogInputExcel(QDialog, Ui_Form):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setupUi(self)
        self.parent = parent
        self.setSizeGripEnabled(True)
        self.SQL = Model_Preferensi()
        self._setup_connections()

    def _setup_connections(self):
        connections = [
            (self.cbo_db.currentIndexChanged, self._on_db_selected),
            (self.cbo_table.currentIndexChanged, self._update_filename),
            (self.btn_no_filter.clicked, self._create_template),
            (self.btn_browse.clicked, self._browse_file),
            (self.cbo_sheet.currentIndexChanged, self._update_sheet_headers),
            (self.insert_btn.clicked, self._execute_insert),
            (self.update_btn.clicked, self._execute_update)
        ]
        for signal, slot in connections:
            signal.connect(slot)

    def show_dialog(self):
        # screen = QApplication.primaryScreen().availableGeometry()
        # self.resize(screen.width() * 0.9, screen.height() * 0.9)
        # self.move((screen.width() - self.width()) // 2, (screen.height() - self.height()) // 2)
        self._fill_db_combobox()

    def _fill_db_combobox(self):
        databases = self.SQL.get_databases()
        for cbo in (self.cbo_db, self.cbo_save_to_db):
            populate_combobox(cbo, databases)

    def _on_db_selected(self):
        self.SQL.set_db(self.cbo_db.currentText())
        tables = ["All"] + self.SQL.get_all_tables()
        populate_combobox(self.cbo_table, tables)
        self._update_filename()

    def _update_filename(self):
        filename = f"{self.cbo_db.currentText()}_{self.cbo_table.currentText()}.xlsx"
        self.line_filename_nofilter.setText(filename)

    def _create_template(self):
        filename = save_as_path(self, self.line_filename_nofilter.text())
        if not filename:
            return
        try:
            success = self._generate_excel_template(filename)
            if success:
                save_value_to_db("LAST_SELECTED_FOLDER", os.path.dirname(filename))
                QMessageBox.information(self.parent, "Sukses", "Template berhasil dibuat")
                open_with_default_app(filename)
        except Exception as e:
            print(f"Error creating template: {str(e)}")

    def _generate_excel_template(self, filename):
        wb = Workbook()
        ws_nav = wb.active
        ws_nav.title = "navigasi"
        ws_nav["A1"] = "MENUJU SHEET"
        ws_nav["A1"].font = Font(bold=True, size=12)

        tables = self.SQL.get_all_tables() if self.cbo_table.currentText() == "All" else [self.cbo_table.currentText()]
        styles = {
            'center': Alignment(horizontal='center', vertical='center'),
            'left': Alignment(horizontal='left', vertical='center', indent=1),
            'font': Font(name='Aptos', size=10),
            'bold': Font(name='Aptos', size=10, bold=True)
        }

        for idx, table in enumerate(tables, 2):
            ws_nav.cell(row=idx, column=1, value=table).hyperlink = f"#'{table}'!A1"
            ws_nav.cell(row=idx, column=1).font = Font(color="0000FF", underline="single")
            ws_nav.cell(row=idx, column=1).alignment = styles['left']

            ws = wb.create_sheet(title=table)
            columns = self.SQL.get_column_names(table)
            data = self.SQL.get_table_data(table) if self.radio_filled.isChecked() else []

            for col, key in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col, value=key)
                cell.font = styles['bold']
                cell.fill = PatternFill(start_color="FFFFCC", fill_type="solid")
                cell.alignment = styles['center']

            ws["A1"].hyperlink = "#'navigasi'!A1"
            ws["A1"].font = Font(underline="single")
            ws["A1"].fill = PatternFill(start_color="4215FF", fill_type="solid")
            ws.freeze_panes = "A2"

            for row, row_data in enumerate(data, 2):
                for col, key in enumerate(columns, 1):
                    cell = ws.cell(row=row, column=col, value=row_data.get(key, ""))
                    cell.alignment = styles['left'] if key in static_values['LEFT_COLUMN'] else styles['center']
                    cell.font = styles['font']

            for col, col_name in enumerate(columns, 1):
                col_letter = get_column_letter(col)
                max_length = max(len(str(col_name)), max((len(str(row_data.get(col_name, ""))) for row_data in data), default=0))
                ws.column_dimensions[col_letter].width = max_length + 2

        wb.save(filename)
        return True

    def _browse_file(self):
        file = open_dialog(parent=self, text_widget=self.line_source)
        if file:
            self.line_namafile.setText(os.path.basename(self.line_source.text()))
            sheets = pd.ExcelFile(self.line_source.text()).sheet_names[1:]
            populate_combobox(self.cbo_sheet, sheets)
        else:
            self.cbo_sheet.clear()
            self.line_namafile.clear()
            self.cbo_key.clear()

    def _update_sheet_headers(self):
        if self.line_source.text():
            df = pd.read_excel(self.line_source.text(), sheet_name=self.cbo_sheet.currentText(), nrows=0)
            populate_combobox(self.cbo_key, df.columns.tolist())

    def _execute_insert(self):
        db_name = self.cbo_save_to_db.currentText()
        con = ConnectDB(db_name)
        con.connect()
        cursor = con.my_cursor
        queries = generate_insert_queries(self.cbo_sheet.currentText(), self.line_source.text())
        if not queries:
            QMessageBox.information(self, "Informasi", "Tidak ada data yang perlu disimpan.")
            return
        self.progress_save.setMaximum(len(queries))
        self.progress_save.setValue(0)

        try:
            insert_queries = [(q, p) for q, p in queries if q.startswith("INSERT")]
            if len(insert_queries) > 1000:
                cursor.executemany(insert_queries[0][0], [p for _, p in insert_queries])
                self.progress_save.setValue(len(insert_queries))
            else:
                for i, (query, params) in enumerate(insert_queries, 1):
                    cursor.execute(query, params)
                    self.progress_save.setValue(i)
            con.my_connector.commit()
            QMessageBox.information(self, "Sukses", "Data berhasil disimpan ke database.")
        except Exception as e:
            con.my_connector.rollback()
            QMessageBox.critical(self, "Error", f"Terjadi kesalahan: {str(e)}")
        finally:
            cursor.close()
            con.close_connection()

    def _execute_update(self):
        db_name = self.cbo_save_to_db.currentText()
        table_name = self.cbo_sheet.currentText()
        key_column = self.cbo_key.currentText()
        source_file = self.line_source.text()

        if not all([db_name, table_name, key_column, source_file]):
            QMessageBox.warning(self, "Peringatan", "Semua field harus diisi.")
            return

        try:
            # Load data from Excel dengan konversi tipe data yang tepat
            df = pd.read_excel(source_file, sheet_name=table_name, dtype=str)  # Baca semua sebagai string dulu
            
            # Konversi ke tipe data yang sesuai
            for col in df.columns:
                if col in static_values['KOLOM_ANGKA']:
                    df[col] = pd.to_numeric(df[col], errors='coerce').astype('Int64')  # Int64 mendukung NaN
                elif col in static_values['KOLOM_FLOAT']:
                    df[col] = pd.to_numeric(df[col], errors='coerce')
                elif col in static_values['KOLOM_TANGGAL']:
                    df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime('%Y-%m-%d')
                elif col in static_values['KOLOM_CURRENCY']:
                    # Hilangkan karakter non-numeric lalu konversi ke float
                    df[col] = df[col].astype(str).str.replace(r'[^\d.]', '', regex=True)
                    df[col] = pd.to_numeric(df[col], errors='coerce')
            
            # Convert NaN values to None (akan menjadi NULL di database)
            df = df.where(pd.notnull(df), None)
            
            if key_column not in df.columns:
                QMessageBox.critical(self, "Error", f"Kolom '{key_column}' tidak ditemukan di sheet.")
                return

            # Connect to database
            con = ConnectDB(db_name)
            con.connect()
            cursor = con.my_cursor

            # Get column names from database table
            columns = self.SQL.get_column_names(table_name)
            valid_columns = [col for col in df.columns if col in columns and col != key_column]

            if not valid_columns:
                QMessageBox.warning(self, "Peringatan", "Tidak ada kolom yang valid untuk diupdate.")
                return

            # Generate update queries
            queries = []
            for _, row in df.iterrows():
                set_clause = ", ".join([f"{col} = %s" for col in valid_columns])
                query = f"UPDATE {table_name} SET {set_clause} WHERE {key_column} = %s"
                
                # Handle parameter dengan tipe data yang tepat
                params = []
                for col in valid_columns:
                    val = row[col]
                    if val is None:
                        params.append(None)
                    elif col in static_values['KOLOM_TANGGAL'] and isinstance(val, str):
                        params.append(val)  # Tanggal sudah dalam format string
                    elif col in static_values['KOLOM_ANGKA'] and isinstance(val, (int, np.integer)):
                        params.append(int(val))
                    elif col in (static_values['KOLOM_FLOAT'] + static_values['KOLOM_CURRENCY']) and isinstance(val, (float, np.floating)):
                        params.append(float(val))
                    else:
                        params.append(str(val) if val is not None else None)
                
                params.append(row[key_column] if pd.notnull(row[key_column]) else None)
                queries.append((query, params))

            # Execute updates
            self.progress_save.setMaximum(len(queries))
            self.progress_save.setValue(0)

            for i, (query, params) in enumerate(queries, 1):
                cursor.execute(query, params)
                self.progress_save.setValue(i)

            con.my_connector.commit()
            QMessageBox.information(self, "Sukses", "Data berhasil diupdate di database.")

        except Exception as e:
            con.my_connector.rollback()
            QMessageBox.critical(self, "Error", f"Terjadi kesalahan: {str(e)}")

        finally:
            cursor.close()
            con.close_connection()