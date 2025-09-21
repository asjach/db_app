from PySide6.QtWidgets import QWidget, QMainWindow, QFileDialog
from utils.fungsi.general_functions import *
from utils.fungsi.functions import read_excel
from ui.ui_page_input_nilai import Ui_Form
from models.model_nilai import Model_Nilai
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
import pandas as pd
from collections import defaultdict
# from utils.static_values import LEFT_COLUMN, RIGHT_COLUMN
import os

class PageInputNilai(QWidget, Ui_Form):
    def __init__(self, parent:QMainWindow):
        super().__init__(parent)
        self.setupUi(self)
        self.parent = parent
        # self.cbo_jenjang = self.parent.cbo_jenjang
        self.cbo_tapel = self.parent.cbo_tapel
        self.SQL = Model_Nilai()
        self.cbo_kegiatan.currentIndexChanged.connect(self.cbo_kegiatan_selected)
        self.btn_new_default_path.clicked.connect(self.browse_save_folder)
        self.open_default_folder.clicked.connect(lambda: open_in_explorer(self.pte_default_path.toPlainText()))
        self.btn_template_nilai.clicked.connect(self.btn_template_nilai_clicked)
        self.btn_template_walas.clicked.connect(self.btn_template_walas_clicked)
        self.btn_template_rekap.clicked.connect(self.btn_template_rekap_clicked)
        self.btn_open_file.clicked.connect(self.btn_open_file_clicked)

        self.radio_nilai_db.clicked.connect(self.show_data_nilai_in_table)
        self.radio_catatan_db.clicked.connect(self.show_data_nilai_in_table)
        self.radio_nilai_catatan_db.clicked.connect(self.show_data_nilai_in_table)
        self.radio_nilai_excel.clicked.connect(self.show_data_nilai_in_table)
        self.btn_browse.clicked.connect(self.browse_btn_clicked)
        self.btn_save.clicked.connect(self.save_btn_clicked)

    def show_page(self): 
        self.fill_kegiatan_cbo()
        self.show_data_nilai_in_table()
        
# CBO KEGIATAN
    def fill_kegiatan_cbo(self):
        data = self.SQL.get_id_kegiatan(
            jenjang=self.parent.str_jenjang,
            tapel=self.cbo_tapel.currentText())
        self.cbo_kegiatan.clear()
        for item in data:
            self.cbo_kegiatan.addItem(item['kegiatan'], userData=item['id'])
        
    def cbo_kegiatan_selected(self):
        self.init_default_folder()
        self.init_file_rekap_nilai()
        self.create_filename()
        self.show_data_nilai_in_table()

# BROWSE FOLDER
    def browse_save_folder(self):
        folder_path = QFileDialog.getExistingDirectory(
            self,
            "Pilih Folder Penyimpanan Template",
            self.pte_default_path.toPlainText())
        if folder_path:
            self.pte_default_path.setPlainText(folder_path)
            self.SQL.update_path(
                jenjang=self.parent.str_jenjang,
                tapel=self.cbo_tapel.currentText(),
                kegiatan=self.cbo_kegiatan.currentText(),
                kolom="path_folder_nilai",
                nilai=self.pte_default_path.toPlainText())

# TEMPLATE NILAI
    def btn_template_nilai_clicked(self):
        jenjang = self.parent.cbo_jenjang.currentText()
        data_siswa = self.get_data_siswa()
        if self.pte_default_path.toPlainText() !=  '':
            if jenjang == 'MI':
                self.template_nilai_mi(data=data_siswa)
            elif jenjang == 'MD':
                self.template_nilai_md(data=data_siswa)
            

# TEMPLATE WALAS
    def btn_template_walas_clicked(self):
        if self.pte_default_path.toPlainText() !=  '':
            self.template_walas(self.get_data_siswa())

# TEMPLATE REKAP
    def btn_template_rekap_clicked(self):
        if self.pte_default_path.toPlainText() !=  '':
            self.template_rekap(self.get_data_siswa())

#   BTN RELOAD
    def show_data_nilai_in_table(self):
        path_rekap = self.pte_excel_path.toPlainText()
        if self.radio_nilai_db.isChecked():
            self.get_rekap_nilai()
        elif self.radio_catatan_db.isChecked():
            self.get_rekap_catatan()
        elif self.radio_nilai_catatan_db.isChecked():
            self.fill_tabel_rekap_nilai_catatan()
        else:
            if path_rekap:
                if os.path.exists(path_rekap):
                    df = pd.read_excel(path_rekap)
                    generate_table(data=df,table=self.input_tbl)
                else:
                    self.input_tbl.clear()
                    self.input_tbl.setRowCount(0)
                    self.input_tbl.setColumnCount(0)
        
    def get_rekap_nilai(self):
        list_kelas = self.parent.str_kelas
        list_tingkat = self.parent.not_quoted_tingkat
        jenjang = self.parent.str_jenjang
        tapel = self.cbo_tapel.currentText()
        tingkat = list_tingkat
        kelas = list_kelas
        kegiatan = self.cbo_kegiatan.currentText()
        data_mapel = self.SQL.get_list_mapel(jenjang, tapel, kegiatan, tingkat, kelas, )
        print(data_mapel)
        if data_mapel:
            mapel_list = [mapel['mapel'] for mapel in data_mapel]
            kolom_mapel = ", ".join([f"MAX(CASE WHEN mapel = '{mapel}' THEN nilai END) AS `{mapel}`" for mapel in mapel_list])
        else:
            kolom_mapel = ""
        data = self.SQL.get_nilai_by_kegiatan(kolom_mapel, jenjang, tapel, tingkat, kelas, kegiatan)
        generate_table(data=data,table=self.input_tbl)

    def get_rekap_catatan(self):
        jenjang = self.parent.str_jenjang
        tapel = self.cbo_tapel.currentText()
        tingkat = self.parent.not_quoted_tingkat
        kelas = self.parent.str_kelas
        kegiatan = self.cbo_kegiatan.currentText()
        data = self.SQL.get_catatan_by_kegiatan(jenjang, tapel, tingkat, kelas, kegiatan)
        generate_table(data=data,table=self.input_tbl)

    def get_rekap_nilai_catatan(self):
        jenjang = self.parent.str_jenjang
        tapel = self.cbo_tapel.currentText()
        tingkat = self.parent.not_quoted_tingkat
        kelas = self.parent.str_kelas
        kegiatan = self.cbo_kegiatan.currentText()
        data_mapel = self.SQL.get_list_mapel(jenjang, tapel, kegiatan, tingkat, kelas, )
        if data_mapel:
            mapel_list = [mapel['mapel'] for mapel in data_mapel]
            kolom_mapel = ", ".join([f"MAX(CASE WHEN mapel = '{mapel}' THEN nilai END) AS `{mapel}`" for mapel in mapel_list])
        else:
            kolom_mapel = ""
        return self.SQL.get_nilai_catatan_by_kegiatan(kolom_mapel, jenjang, tapel, tingkat, kelas, kegiatan)
        
    def fill_tabel_rekap_nilai_catatan(self):
        data = self.get_rekap_nilai_catatan()
        generate_table(data=data,table=self.input_tbl)

#   BTN OPEN FILE REKAP           
    def btn_open_file_clicked(self):
        path = self.pte_excel_path.toPlainText()
        if path != '':
            open_with_default_app(path)
        else:
            self.browse_btn_clicked()

#   BTN BROWSE/MEMILIH FILE REKAP
    def browse_btn_clicked(self):
        open_dialog(self, self.pte_excel_path)
        if self.pte_excel_path.toPlainText() !='':
            self.SQL.update_path(
                jenjang=self.parent.str_jenjang,
                tapel = self.cbo_tapel.currentText(),
                kegiatan=self.cbo_kegiatan.currentText(),
                kolom="path_file_rekap",
                nilai=self.pte_excel_path.toPlainText())

#   BTN SAVE CLICKED
    def save_btn_clicked(self):
        if self.pte_excel_path.toPlainText() != '':
            data = read_excel(self.pte_excel_path.toPlainText())
            self.save_operation(data)
        else:
            print("Tidak ada file excel yang dipilih")

    
#   HELPERS METHOD
    def get_data_siswa(self):
        data_siswa = self.SQL.data_siswa(
            jenjang=self.parent.str_jenjang, 
            tapel= self.cbo_tapel.currentText(), 
            kegiatan= self.cbo_kegiatan.currentText()
        )
        if not data_siswa:
            print("Tidak ada data siswa untuk kegiatan ini.")
            return None
        return data_siswa
    
    def create_filename(self):
        kegiatan = self.cbo_kegiatan.currentText()
        jenjang = self.parent.str_jenjang
        tapel = self.cbo_tapel.currentText()
        filename_rekap = f"Rekap Nilai {kegiatan} {jenjang} {tapel}.xlsx"
        self.line_rekap.setText(filename_rekap)
        
    def init_default_folder(self):
        default_path = self.SQL.get_path(
            jenjang=self.parent.str_jenjang,
            tapel = self.cbo_tapel.currentText(),
            kegiatan=self.cbo_kegiatan.currentText(),
            kolom="path_folder_nilai")
        if default_path:
            self.pte_default_path.setPlainText(default_path[0]['path_folder_nilai'])
        else:
            self.pte_default_path.clear()

    def init_file_rekap_nilai(self):
        file_rekap_nilai = self.SQL.get_path(
            jenjang=self.parent.str_jenjang,
            tapel = self.cbo_tapel.currentText(),
            kegiatan=self.cbo_kegiatan.currentText(),
            kolom="path_file_rekap")
        if file_rekap_nilai:
            self.pte_excel_path.setPlainText(file_rekap_nilai[0]['path_file_rekap'])
        else:
            self.pte_excel_path.clear()

    def template_nilai_md(self, data):
        # Extract form data
        jenjang = self.parent.str_jenjang
        tapel = self.cbo_tapel.currentText()
        kegiatan = self.cbo_kegiatan.currentText()

        # Define columns
        base_columns = ["no_urut", "nis_lokal", "nama_lengkap", "kelas"]
        additional_columns = [f"nh{i}" for i in range(1, 7)] + ["nrh", "kehadiran", "tulis", "lisan", "nilai"]
        all_columns = base_columns + additional_columns

        # Create output directory
        folder_output = os.path.join(os.path.normpath(self.pte_default_path.toPlainText()), 'blanko nilai')
        os.makedirs(folder_output, exist_ok=True)

        # Get class list
        kelas = self.SQL.get_kelas(jenjang, tapel)
        list_kelas = [item['kelas'] for item in kelas]

        # Define styles
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin'))
        default_font = Font(name='Aptos', size=11)
        header_font = Font(name='Aptos', size=11, bold=True)
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        red_rule = CellIsRule(operator="greaterThan", formula=["100"], fill=red_fill)

        # Process data
        df = pd.DataFrame(data, columns=base_columns)
        for col in additional_columns:
            df[col] = None

        for kls in list_kelas:
            df_kelas = df[df['kelas'] == kls].copy()
            filename = os.path.join(folder_output, f'{kegiatan} {kls} {jenjang}.xlsx')
            wb = Workbook()
            ws = wb.active

            # Set headers
            for col_idx, col_name in enumerate(all_columns, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = col_name
                cell.border = thin_border
                cell.fill = yellow_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = header_font

            # Populate data
            for row_idx, row_data in enumerate(df_kelas.values, start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.border = thin_border
                    cell.font = default_font
                    col_name = all_columns[col_idx - 1]
                    cell.alignment = Alignment(
                        horizontal='left' if col_name in static_values['LEFT_COLUMN'] else 
                        'right' if col_name in static_values['RIGHT_COLUMN'] else 'center',
                        vertical='center'
                    )

            # Set formulas
            for row in range(2, len(df_kelas) + 2):
                cell_nrh = ws[f"K{row}"]
                cell_nilai = ws[f"O{row}"]
                cell_nrh.value = f"=AVERAGE(E{row}:J{row})"
                cell_nilai.value = f"=ROUNDUP(AVERAGE(K{row}:N{row}),0)"
                cell_nrh.border = thin_border
                cell_nilai.border = thin_border
                cell_nrh.font = default_font
                cell_nilai.font = default_font
                cell_nrh.alignment = Alignment(
                    horizontal='left' if 'nrh' in static_values['LEFT_COLUMN'] else 
                    'right' if 'nrh' in static_values['RIGHT_COLUMN'] else 'center',
                    vertical='center'
                )
                cell_nilai.alignment = Alignment (
                    horizontal='left' if 'nilai' in static_values['LEFT_COLUMN'] else 
                    'right' if 'nilai' in static_values['RIGHT_COLUMN'] else 'center',
                    vertical='center'
                )

            # Apply conditional formatting
            for col_name in additional_columns:
                col_idx = all_columns.index(col_name) + 1
                col_letter = get_column_letter(col_idx)
                ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{len(df_kelas) + 1}", red_rule)

            # Adjust column widths
            for col_idx, col_name in enumerate(all_columns, start=1):
                if col_name in ['nrh', 'nilai']:
                    continue
                max_length = max(
                    len(str(col_name)),
                    *(len(str(ws.cell(row=row_idx, column=col_idx).value or '')) 
                    for row_idx in range(1, len(df_kelas) + 2))
                )
                ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 3

            wb.save(filename)
        
        open_in_explorer(folder_output)

    def template_nilai_mi(self, data):
        # Extract form data
        jenjang = self.parent.str_jenjang
        tapel = self.cbo_tapel.currentText()
        kegiatan = self.cbo_kegiatan.currentText()

        # Define columns
        base_columns = ["no_urut", "nis_lokal", "nama_lengkap", "kelas"]
        additional_columns = [f"nh{i}" for i in range(1, 9)] + ["nrh", "tulis", "lisan", "nilai"]
        all_columns = base_columns + additional_columns

        # Create output directory
        folder_output = os.path.join(os.path.normpath(self.pte_default_path.toPlainText()), 'blanko nilai')
        os.makedirs(folder_output, exist_ok=True)

        # Get class list
        kelas = self.SQL.get_kelas(jenjang, tapel)
        list_kelas = [item['kelas'] for item in kelas]

        # Define styles
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin'))
        default_font = Font(name='Aptos', size=11)
        header_font = Font(name='Aptos', size=11, bold=True)
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        red_rule = CellIsRule(operator="greaterThan", formula=["100"], fill=red_fill)

        # Process data
        df = pd.DataFrame(data, columns=base_columns)
        for col in additional_columns:
            df[col] = None

        for kls in list_kelas:
            df_kelas = df[df['kelas'] == kls].copy()
            filename = os.path.join(folder_output, f'{kegiatan} {kls} {jenjang}.xlsx')
            wb = Workbook()
            ws = wb.active

            # Set headers
            for col_idx, col_name in enumerate(all_columns, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = col_name
                cell.border = thin_border
                cell.fill = yellow_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = header_font

            # Populate data
            for row_idx, row_data in enumerate(df_kelas.values, start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.border = thin_border
                    cell.font = default_font
                    col_name = all_columns[col_idx - 1]
                    cell.alignment = Alignment(
                        horizontal='left' if col_name in static_values['LEFT_COLUMN'] else 
                        'right' if col_name in static_values['RIGHT_COLUMN'] else 'center',
                        vertical='center'
                    )

            # Set formulas
            for row in range(2, len(df_kelas) + 2):
                cell_nrh = ws[f"M{row}"]
                cell_nilai = ws[f"P{row}"]
                cell_nrh.value = f"=AVERAGE(E{row}:L{row})"
                cell_nilai.value = f"=ROUNDUP((M{row} * 0.6) + (AVERAGE(N{row}:O{row}) * 0.4),0)"
                cell_nrh.border = thin_border
                cell_nilai.border = thin_border
                cell_nrh.font = default_font
                cell_nilai.font = header_font
                cell_nrh.alignment = Alignment(
                    horizontal='left' if 'nrh' in static_values['LEFT_COLUMN'] else 
                    'right' if 'nrh' in static_values['RIGHT_COLUMN'] else 'center',
                    vertical='center'
                )
                cell_nilai.alignment = Alignment(
                    horizontal='left' if 'nilai' in static_values['LEFT_COLUMN'] else 
                    'right' if 'nilai' in static_values['RIGHT_COLUMN'] else 'center',
                    vertical='center'
                )

            # Apply conditional formatting
            for col_name in additional_columns:
                col_idx = all_columns.index(col_name) + 1
                col_letter = get_column_letter(col_idx)
                ws.conditional_formatting.add(f"{col_letter}2:{col_letter}{len(df_kelas) + 1}", red_rule)

            # Adjust column widths
            for col_idx, col_name in enumerate(all_columns, start=1):
                if col_name in ['nrh', 'nilai']:
                    continue
                max_length = max(
                    len(str(col_name)),
                    *(len(str(ws.cell(row=row_idx, column=col_idx).value or '')) 
                    for row_idx in range(1, len(df_kelas) + 2))
                )
                ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 3

            wb.save(filename)
        
        open_in_explorer(folder_output)

    def template_walas(self, data):
        # Extract form data
        jenjang = self.parent.str_jenjang
        tapel = self.cbo_tapel.currentText()
        kegiatan = self.cbo_kegiatan.currentText()

        # Define columns
        base_columns = ["no_urut", "nis_lokal", "nama_lengkap", "kelas"]
        additional_columns = ['sakit', 'ijin', 'alpa', 'catatan_walas']
        if kegiatan == 'PAT':
            additional_columns.append('status_naik')
        all_columns = base_columns + additional_columns

        # Create output directory
        folder_output = os.path.join(os.path.normpath(self.pte_default_path.toPlainText()), 'Blanko Walas')
        os.makedirs(folder_output, exist_ok=True)

        # Get class list
        kelas = self.SQL.get_kelas(jenjang, tapel)
        list_kelas = [item['kelas'] for item in kelas]

        # Define styles
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin'))
        default_font = Font(name='Aptos', size=11)
        header_font = Font(name='Aptos', size=11, bold=True)
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # Process data
        df = pd.DataFrame(data, columns=base_columns)
        for col in additional_columns:
            df[col] = None

        for kls in list_kelas:
            df_kelas = df[df['kelas'] == kls].copy()
            filename = os.path.join(folder_output, f'Catatan_Walas_{kegiatan} {kls} {jenjang}.xlsx')
            wb = Workbook()
            ws = wb.active

            # Set headers with yellow fill
            for col_idx, col_name in enumerate(all_columns, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = col_name
                cell.border = thin_border
                cell.fill = yellow_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = header_font

            # Populate data
            for row_idx, row_data in enumerate(df_kelas.values, start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.border = thin_border
                    cell.font = default_font
                    col_name = all_columns[col_idx - 1]
                    cell.alignment = Alignment(
                        horizontal='left' if col_name in static_values['LEFT_COLUMN'] else 
                        'right' if col_name in static_values['RIGHT_COLUMN'] else 'center',
                        vertical='center'
                    )

            # Add data validation for status_naik if applicable
            if kegiatan == 'PAT':
                status_col_idx = all_columns.index('status_naik') + 1
                status_col_letter = get_column_letter(status_col_idx)
                dv = DataValidation(type="list", formula1='"Naik,Tidak Naik"', allow_blank=True)
                dv.add(f"{status_col_letter}2:{status_col_letter}{len(df_kelas) + 1}")
                ws.add_data_validation(dv)

            # Adjust column widths
            for col_idx, col_name in enumerate(all_columns, start=1):
                max_length = max(
                    len(str(col_name)),
                    *(len(str(ws.cell(row=row_idx, column=col_idx).value or '')) 
                    for row_idx in range(1, len(df_kelas) + 2))
                )
                ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

            wb.save(filename)
        
        open_in_explorer(folder_output)

    def template_rekap(self, data):
        namafile = self.line_rekap.text()
        save_path = self.pte_default_path.toPlainText()
        filename = os.path.join(save_path, namafile)
        jenjang = self.parent.str_jenjang
        tapel = self.cbo_tapel.currentText()
        kegiatan = self.cbo_kegiatan.currentText()
        kolom = ["id_kelas", "id_kegiatan", "no_urut", "id_peserta", "nis_lokal", "nama_lengkap", "kelas"]
        list_of_dict_mapel = self.SQL.get_list_mapel(jenjang, tapel, kegiatan, '', '')
        merge_mapel = list(dict.fromkeys(d['mapel'] for d in list_of_dict_mapel))
        kolom_tambahan = merge_mapel + ['rata_rata', 'jumlah', 'ranking', 'sakit', 'ijin', 'alpa', 'ranking', 'catatan_walas']
        if kegiatan in ['PAT']:
            kolom_tambahan = kolom_tambahan + ['status_naik']
        df = pd.DataFrame(data, columns=kolom)
        for col in kolom_tambahan:
            df[col] = ""
        df.to_excel(filename, index=False)
        wb = load_workbook(filename)
        ws = wb.active
        all_columns = kolom + kolom_tambahan
        kolom_pelajaran_awal = get_column_letter(all_columns.index('kelas')+2)
        end_kolom = get_column_letter(all_columns.index('rata_rata'))
        kolom_jumlah = get_column_letter(all_columns.index('jumlah')+1)
        kolom_rata_rata = get_column_letter(all_columns.index('rata_rata')+1)
        kolom_ranking = get_column_letter(all_columns.index('ranking')+1)
        kelas_ranges = defaultdict(list)
        for row in range(2, len(df) + 2):
            ws[f"{kolom_rata_rata}{row}"] = f"=AVERAGE({kolom_pelajaran_awal}{row}:{end_kolom}{row})"
            ws[f"{kolom_jumlah}{row}"] = f"=SUM({kolom_pelajaran_awal}{row}:{end_kolom}{row})"
        for row_idx, row in enumerate(df.itertuples(), start=2):
            kelas_ranges[row.kelas].append(row_idx)
        for kelas, rows in kelas_ranges.items():
            start_row = rows[0]
            end_row = rows[-1]
            for row in rows:
                ws[f"{kolom_ranking}{row}"] = f"=RANK({kolom_jumlah}{row},${kolom_jumlah}${start_row}:${kolom_jumlah}${end_row})"
        exclude_column = ['rata_rata', 'jumlah', 'ranking']
        all_columns = kolom + kolom_tambahan
        for col_idx, col_name in enumerate(all_columns, start=1):
            if col_name in exclude_column:
                continue
            max_length = max(
                len(str(col_name)),  
                *(len(str(ws.cell(row=row_idx, column=col_idx).value)) for row_idx in range(1, len(df) + 2)))
            ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2
        wb.save(filename)
        open_in_explorer(filename)

    def save_operation(self, data):
        MAPEL_NILAI = self.SQL.all_mapel()
        KEGIATAN_FIELDS = ['no_urut', 'sakit', 'ijin', 'alpa', 'catatan_walas', 'ranking', 'status_naik']
        nilai_keys = [(row['id_peserta'], mapel) for row in data for mapel in MAPEL_NILAI if mapel in row]
        existing_nilai_dict = self.SQL.cek_nilai_bulk(nilai_keys)
        peserta_keys = [row['id_peserta'] for row in data]
        existing_peserta_dict = self.SQL.cek_peserta_bulk(peserta_keys)
        insert_nilai_data = []
        update_nilai_data = []
        insert_peserta_data = []
        update_peserta_data = []
        for row in data:
            id_peserta = int(row['id_peserta'])
            id_kelas = int(row['id_kelas'])
            id_kegiatan = int(row['id_kegiatan'])
            mapel_keys = [key for key in row.keys() if key in MAPEL_NILAI]
            for nama_mapel in mapel_keys:
                nilai = row[nama_mapel]
                if isinstance(nilai, str):
                    nilai = nilai.strip()
                if not nilai or nilai == '':
                    nilai = None
                else:
                    try:
                        nilai = float(str(nilai))
                    except ValueError:
                        nilai = None
                nilai_key = (id_peserta, nama_mapel)
                if nilai_key in existing_nilai_dict:
                    id, nilai_db = existing_nilai_dict[nilai_key]
                    if nilai != nilai_db:
                        update_nilai_data.append((nilai, id))
                else:
                    insert_nilai_data.append((id_peserta, nama_mapel, nilai))
            
            peserta_data = {field: row.get(field) for field in KEGIATAN_FIELDS}
            peserta_key = id_peserta
            
            if peserta_key in existing_peserta_dict:
                existing_data = existing_peserta_dict[peserta_key]
                if any(peserta_data[field] != existing_data[field] for field in KEGIATAN_FIELDS):
                    update_peserta_data.append((
                        peserta_data['no_urut'], 
                        peserta_data['sakit'], 
                        peserta_data['ijin'],
                        peserta_data['alpa'], 
                        peserta_data['catatan_walas'], 
                        peserta_data['ranking'],
                        peserta_data['status_naik'],
                        id_peserta
                    ))
            else:
                insert_peserta_data.append((
                    id_peserta, id_kelas, id_kegiatan,
                    peserta_data['no_urut'], 
                    peserta_data['sakit'], 
                    peserta_data['ijin'],
                    peserta_data['alpa'], 
                    peserta_data['catatan_walas'],
                    peserta_data['ranking'],
                    peserta_data['status_naik'],
                ))
        pesan_insert_nilai = ''
        pesan_update_nilai = ''
        pesan_insert_peserta = ''
        pesan_update_peserta = ''
        if insert_nilai_data:
            self.SQL.insert_nilai_bulk(insert_nilai_data)
            pesan_insert_nilai = f"Bulk inserted {len(insert_nilai_data)} records to nilai_angka"
        if update_nilai_data:
            self.SQL.update_nilai_bulk(update_nilai_data)
            pesan_update_nilai = f"Bulk updated {len(update_nilai_data)} records in nilai_angka"
        if insert_peserta_data:
            self.SQL.insert_peserta_bulk(insert_peserta_data)
            pesan_insert_peserta = f"Bulk inserted {len(insert_peserta_data)} records to kegiatan_peserta"
        if update_peserta_data:
            self.SQL.update_peserta_bulk(update_peserta_data)
            pesan_update_peserta = f"Bulk updated {len(update_peserta_data)} records in kegiatan_peserta"
        pesan_sukses(
            judul="Berhasil Insert Data",
            pesan=f'{pesan_insert_nilai}\n{pesan_update_nilai}\n{pesan_insert_peserta}\n{pesan_update_peserta}'
        )
        self.show_page()