import json
from typing import Optional, Dict, List, Any, Union
from PySide6.QtWidgets import QWidget, QMainWindow, QMessageBox, QVBoxLayout
from PySide6.QtCore import QTimer, Signal, QObject
from reportlab.lib.pagesizes import A4, GOV_LEGAL
from ui.ui_page_rekap_nilai import Ui_Form
from utils.fungsi.general_functions import generate_table, table_selected, save_pdf, print_with_foxit, format_cell_data, header_for_table
from models.model_nilai import Model_Nilai
from template.rekap_nilai import TemplateRekapNilai
from scripts.widgets.dokumen_viewer import DokumenViewer
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from PySide6.QtWidgets import QFileDialog

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class PageRekapNilaiSignals(QObject):
    """Signals for PageRekapNilai to enable thread-safe operations"""
    error_occurred = Signal(str, str)  # title, message
    pdf_ready = Signal(bytes)


class PageRekapNilai(Ui_Form, QWidget):
    """
    A widget for displaying and managing grade summaries with PDF generation capabilities.
    
    Features:
    - Display grade summaries for classes or top students
    - Customize PDF output settings
    - Save or print generated PDFs
    """
    
    # Constants
    PDF_UPDATE_DELAY_MS = 200
    DEFAULT_KKM = 60
    DEFAULT_PAPER_SIZE = 'F4'
    DEFAULT_ORIENTATION = 'Portrait'
    DEFAULT_MARGIN = 1.5
    DEFAULT_FONT_SIZE = 11
    DEFAULT_COL_NAMA_WIDTH = 6
    DEFAULT_SUBJECT_COL_WIDTH = 0.6
    DEFAULT_COL_AYAH = 4
    DEFAULT_COL_IBU = 3
    DEFAULT_COL_ALAMAT = 3
    DEFAULT_COL_JML = 1
    DEFAULT_COL_RT = 1
    DEFAULT_ROW_HEIGHT = 0.6
    DEFAULT_ROW_HEIGHT_1 = 1
    DEFAULT_ROW_HEIGHT_3 = 0.8
    DEFAULT_ROW_HEIGHT_10 = 0.6
    DEFAULT_AYAH = True
    DEFAULT_IBU = True
    DEFAULT_ALAMAT = True
    DEFAULT_FOLDER = ''

    def __init__(self, parent: Optional[QMainWindow] = None) -> None:
        """Initialize the grade summary page.
        
        Args:
            parent: Optional parent QMainWindow
        """
        super().__init__(parent)
        self.setupUi(self)
        self.parent = parent
        # Initialize components
        self._init_components()
        self._init_variables()
        self._setup_pdf_timer()
        self._setup_database()
        self._setup_viewer()
        self._connect_signals()
        # Setup default settings
        self.DEFAULT_SETTING = self._create_default_settings()

    def _init_components(self):
        """Initialize UI components from parent"""
        if not self.parent:
            raise ValueError("Parent widget is required for initialization")
            
        # self.cbo_jenjang = self.parent.cbo_jenjang
        # self.cbo_tapel = self.parent.cbo_tapel
        # self.cbo_tingkat = self.parent.cbo_tingkat
        # self.cbo_kelas = self.parent.cbo_kelas
        
        # Create signals object
        self.signals = PageRekapNilaiSignals()
        self.signals.error_occurred.connect(self._show_error)
        self.signals.pdf_ready.connect(self._handle_pdf_ready)

    def _init_variables(self):
        """Initialize class variables"""
        self.id_kelas = None
        self.id_kegiatan = None
        self.pdf_data = None
        self.wali_kelas = None
        self.current_settings = {}
        self.gabungan = []

    def _setup_pdf_timer(self):
        """Setup timer for PDF updates"""
        self.pdf_update = QTimer(self)
        self.pdf_update.setSingleShot(True)
        self.pdf_update.timeout.connect(self.update_pdf)

    def _setup_database(self):
        """Initialize database connection"""
        self.SQL = Model_Nilai()
        
    def _setup_viewer(self):
        """Setup PDF viewer"""
        self.viewer = DokumenViewer()
        if not hasattr(self, 'viewer_layout'):
            self.viewer_layout = QVBoxLayout()
            self.setLayout(self.viewer_layout)
        self.viewer_layout.addWidget(self.viewer)

    def _connect_signals(self):
        """Connect signals to slots"""
        # Table selection signals
        self.kegiatan_tbl.itemSelectionChanged.connect(self.kegiatan_tbl_selected)
        self.kelas_tbl.itemSelectionChanged.connect(self.kelas_tbl_selected)
        
        # Button signals
        self.btn_print.clicked.connect(self.print_pdf)
        self.btn_excel.clicked.connect(self._export_ranking_to_excel)
        self.btn_save.clicked.connect(self.save_pdf)
        
        # Combo and spin box signals
        combo = [self.cbo_kertas, self.cbo_orientasi]
        spin = [
            self.margin_left_spin, self.margin_top_spin, self.margin_right_spin,
            self.margin_bottom_spin, self.tinggi_baris_spin, self.kolom_pelajaran_spin,
            self.kolom_nama_spin, self.nilai_merah_spin
        ]
        self.gabungan = combo + spin
        
        for ctl in combo:
            ctl.currentIndexChanged.connect(self.opsi_selected)
        for ctl in spin:
            ctl.valueChanged.connect(self.opsi_selected)
        
        # Radio button signals
        self.pertama_radio.clicked.connect(self.opsi_selected)
        self.tiga_besar_radio.clicked.connect(self.opsi_selected)
        self.perkelas_radio.clicked.connect(self.opsi_selected)
        self.lengkap_radio.clicked.connect(self.opsi_selected)
        self.singkat_radio.clicked.connect(self.opsi_selected)

    def _create_default_settings(self) -> Dict[str, Any]:
        """Create default settings dictionary"""
        return {
            'kertas': self.DEFAULT_PAPER_SIZE,
            'orientasi': self.DEFAULT_ORIENTATION,
            'margin_left': self.DEFAULT_MARGIN,
            'margin_top': self.DEFAULT_MARGIN,
            'margin_right': self.DEFAULT_MARGIN,
            'margin_bottom': self.DEFAULT_MARGIN,
            'folder_rekap': self.DEFAULT_FOLDER,
            'font_size': self.DEFAULT_FONT_SIZE,
            'kkm': self.DEFAULT_KKM,
            'kolom_nama': self.DEFAULT_COL_NAMA_WIDTH,
            'kolom_pelajaran': self.DEFAULT_SUBJECT_COL_WIDTH,
            'ayah': self.DEFAULT_AYAH,
            'ibu':self.DEFAULT_IBU,
            'alamat':self.DEFAULT_ALAMAT,
            'kolom_ayah': self.DEFAULT_COL_AYAH,
            'kolom_ibu': self.DEFAULT_COL_IBU,
            'kolom_alamat':self.DEFAULT_COL_ALAMAT,
            'tinggi_baris': self.DEFAULT_ROW_HEIGHT,
            'tinggi_baris_1' : self.DEFAULT_ROW_HEIGHT_1,
            'tinggi_baris_3' : self.DEFAULT_ROW_HEIGHT_3,
            'tinggi_baris_10' : self.DEFAULT_ROW_HEIGHT_10
        }

    def show_page(self):
        """Show the page and populate activities table"""
        try:
            self.fill_kegiatan_tbl()
        except Exception as e:
            self._show_error("Gagal memuat data kegiatan", str(e))
            logger.error(f"Failed to show page: {str(e)}")

    def fill_kegiatan_tbl(self):
        data, nama_kolom = self.SQL.get_kegiatan_riwayat(self.parent.cbo_tapel.currentText())   
        generate_table(
                data=data,
                table=self.kegiatan_tbl,
                hidden_column=[0,5],
                stretch_column=4,
            )

    def kegiatan_tbl_selected(self):
        """Handle selection change in activities table"""
        try:
            # Dapatkan data yang dipilih
            selected_data = table_selected(self.kegiatan_tbl, self, self.parent)
            
            # Pastikan ada data yang dipilih
            if not selected_data or not isinstance(selected_data, list) or len(selected_data) == 0:
                return
                
            # Ambil data pertama (jika multi-select, kita ambil yang pertama)
            data = selected_data[0]
            
            # Simpan id_kegiatan untuk referensi nanti
            self.id_kegiatan = data.get('id')
            if not self.id_kegiatan:
                return
                
            # Pastikan atribut yang diperlukan sudah tersedia
            # (karena set_attributes_values sudah dijalankan oleh table_selected)
            required_attrs = ['jenjang', 'tapel', 'kegiatan']
            if not all(hasattr(self, attr) for attr in required_attrs):
                logger.warning("Required attributes not set")
                return
                
            # Isi tabel kelas
            self.fill_kelas_tbl(
                jenjang=self.jenjang,
                tapel=self.tapel,
                id_kegiatan=self.id_kegiatan
            )
            
        except Exception as e:
            self._show_error("Error memilih kegiatan", str(e))
            logger.error(f"Failed to handle kegiatan selection: {str(e)}")

    def fill_kelas_tbl(self, jenjang: str, tapel: str, id_kegiatan: int):
        """Populate class table with data"""
        try:
            if not all([jenjang, tapel, id_kegiatan]):
                raise ValueError("Parameter tidak lengkap")
                
            data = self.SQL.get_kelas_riwayat_with_peserta(jenjang, tapel, id_kegiatan)
            if not data:
                logger.warning(f"No class data found for jenjang={jenjang}, tapel={tapel}, id_kegiatan={id_kegiatan}")
                return
                
            generate_table(
                data=data,
                table=self.kelas_tbl,
                hidden_column=[0, 1, 2, 3, 5],
                stretch_column=6
            )
        except Exception as e:
            logger.error(f"Failed to fill kelas table: {str(e)}")
            raise

    def kelas_tbl_selected(self):
        """Handle selection change in class table"""
        try:
            # Dapatkan data yang dipilih
            selected_data = table_selected(self.kelas_tbl, self, self.parent)
            
            # Pastikan ada data yang dipilih
            if not selected_data or not isinstance(selected_data, list) or len(selected_data) == 0:
                return
                
            # Ambil data pertama
            data = selected_data[0]
            
            # Simpan id_kelas untuk referensi nanti
            self.id_kelas = data.get('id')
            if not self.id_kelas:
                return
                
            # Pastikan atribut yang diperlukan sudah tersedia
            required_attrs = ['jenjang', 'tapel', 'tingkat', 'kelas', 'nama_lengkap']
            if not all(hasattr(self, attr) for attr in required_attrs):
                logger.warning("Required attributes not set")
                return
                
            # Load setting dan tampilkan PDF
            self.load_setting_rekap(self.id_kelas)
            self.show_pdf()
            
        except Exception as e:
            self._show_error("Error memilih kelas", str(e))
            logger.error(f"Failed to handle kelas selection: {str(e)}")

    def show_pdf(self):
        """Generate and display PDF based on current settings"""
        try:
            if not self.id_kegiatan:
                logger.warning("No kegiatan selected, cannot show PDF")
                return
                
            # Get paper and orientation settings
            paper_size = GOV_LEGAL if self.cbo_kertas.currentText() == 'F4' else A4
            orientation = "Portrait" if self.cbo_orientasi.currentText() == 'Portrait' else "Landscape"
            
            if self.perkelas_radio.isChecked():
                self._generate_class_pdf(paper_size, orientation)
            else:
                self._generate_ranking_pdf(paper_size, orientation)
                
        except Exception as e:
            self._show_error("Error membuat PDF", str(e))
            logger.error(f"Failed to generate PDF: {str(e)}")

    def _generate_class_pdf(self, paper_size, orientation):
        """Generate PDF for class summary"""
        data_nilai = self.get_rekap_nilai()
        if not data_nilai:
            logger.warning("No grade data available for class PDF")
            return
            
        self.wali_kelas = getattr(self, 'nama_lengkap', '')
        data = self.convert_mysql_to_reportlab(data_nilai)
        
        # Calculate class averages
        average_row = self._calculate_average_row(data[1:])
        if average_row:
            data.append(average_row)
            
        if data:
            template = TemplateRekapNilai(self)
            self.pdf_data = template.build_pdf(
                top=self.margin_top_spin.value(),
                left=self.margin_left_spin.value(),
                bottom=self.margin_bottom_spin.value(),
                right=self.margin_right_spin.value(),
                orientasi=orientation,
                kertas=paper_size,
                data_detail=data
            )
            self.viewer.loadPDF(self.pdf_data)
        else:
            self.pdf_data = None
            self.viewer.close_file() 

    def _generate_ranking_pdf(self, paper_size, orientation):
        """Generate PDF for student rankings"""
        opsi_peringkat = self._opsi_peringkat_selected()
        template = TemplateRekapNilai(self)
        
        data_peringkat = self.SQL.get_daftar_peringkat(
            id_kegiatan=self.id_kegiatan, 
            ayah=self.ayah_radio.isChecked(),
            ibu=self.ibu_radio.isChecked(),
            alamat=self.alamat_radio.isChecked() ,
            opsi=opsi_peringkat
            )
        if not data_peringkat:
            logger.warning(f"No ranking data found for id_kegiatan={self.id_kegiatan}")
            return
            
        data_peringkat = self.convert_mysql_to_reportlab(data_peringkat)
        self.pdf_data = template.build_pdf_peringkat(
            top=self.margin_top_spin.value(),
            left=self.margin_left_spin.value(),
            bottom=self.margin_bottom_spin.value(),
            right=self.margin_right_spin.value(),
            orientasi=orientation,
            kertas=paper_size,
            data_peringkat=data_peringkat,
            opsi_peringkat=opsi_peringkat
        )
        self.viewer.loadPDF(self.pdf_data)

    def _export_ranking_to_excel(self):
        """Export student rankings to Excel using data langsung dari MySQL"""
        opsi_peringkat = self._opsi_peringkat_selected()
        
        # Get complete data (with headers) from MySQL
        data_peringkat = self.SQL.get_daftar_peringkat(
            id_kegiatan=self.id_kegiatan, 
            ayah=self.ayah_radio.isChecked(),
            ibu=self.ibu_radio.isChecked(),
            alamat=self.alamat_radio.isChecked() ,
            opsi=opsi_peringkat
            )
        if not data_peringkat:
            logger.warning(f"No ranking data found for id_kegiatan={self.id_kegiatan}")
            return
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Ranking Siswa"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
        
        # Asumsi data_peringkat adalah list of dicts dengan struktur:
        # [{'column1': value1, 'column2': value2, ...}, ...]
        
        # Jika data sudah include header dalam bentuk key dictionary
        if isinstance(data_peringkat[0], dict):
            # Write headers from dictionary keys
            headers = list(data_peringkat[0].keys())
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=str(header))
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Write data rows
            for row_num, row_data in enumerate(data_peringkat, 2):
                for col_num, key in enumerate(headers, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=str(row_data.get(key, '')))
                    cell.border = thin_border
        
        # Jika data adalah list of lists dengan row pertama sebagai header
        elif isinstance(data_peringkat[0], (list, tuple)):
            # Write header row
            headers = data_peringkat[0]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=str(header))
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Write data rows (skip first row if it's header)
            for row_num, row_data in enumerate(data_peringkat[1:], 2):
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=str(value))
                    cell.border = thin_border
        
        # Auto adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save file
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Simpan File Excel",
            "",
            "Excel Files (*.xlsx)"
        )
        
        if file_path:
            if not file_path.endswith('.xlsx'):
                file_path += '.xlsx'
            
            wb.save(file_path)
            logger.info(f"File Excel berhasil disimpan di: {file_path}")  


    def _opsi_peringkat_selected(self):
       if self.pertama_radio.isChecked(): return 1
       elif self.tiga_besar_radio.isChecked(): return 3
       elif self.sepuluh_besar_radio.isChecked(): return 10
       else: return 0

    def _calculate_average_row(self, data_rows: List[List]) -> List[str]:
        """Calculate average scores row"""
        try:
            nilai_pelajaran = [[self.safe_float(x) for x in row[3:-1]] for row in data_rows]
            rata_rata_pelajaran = [int(round(sum(x) / len(x), 0)) for x in zip(*nilai_pelajaran)]
            return ["RATA-RATA KELAS"] + [""] * 2 + [str(nilai) for nilai in rata_rata_pelajaran]
        except Exception as e:
            logger.error(f"Failed to calculate averages: {str(e)}")
            return []

    def update_pdf(self):
        """Update PDF display and settings"""
        try:
            self.show_pdf()
            self.update_setting()
        except Exception as e:
            logger.error(f"Failed to update PDF: {str(e)}")

    def opsi_selected(self):
        """Handle option selection changes"""
        self.pdf_update.start(self.PDF_UPDATE_DELAY_MS)

    def get_rekap_nilai(self) -> List[Dict]:
        """Get grade summary data for current selection"""
        try:
            if not all([
                getattr(self, 'jenjang', None),
                getattr(self, 'tapel', None),
                getattr(self, 'tingkat', None),
                getattr(self, 'kelas', None),
                getattr(self, 'kegiatan', None)
            ]):
                logger.warning("Incomplete selection for grade summary")
                return []
                
            data_mapel = self.SQL.get_list_mapel(
                self.jenjang, 
                self.tapel, 
                self.kegiatan, 
                self.tingkat, 
                self.kelas
            )
            
            if not data_mapel:
                logger.warning("No subjects found for selected class")
                return []
                
            mapel_list = [mapel['mapel'] for mapel in data_mapel]
            kolom_mapel = ", ".join(
                f"MAX(CASE WHEN mapel = '{mapel}' THEN nilai END) AS `{mapel}`" 
                for mapel in mapel_list
            )
            nama = 'lengkap' if self.lengkap_radio.isChecked() else 'ringkas'
            return self.SQL.get_nilai_by_kegiatan(
                kolom_mapel, 
                self.jenjang, 
                self.tapel, 
                self.tingkat, 
                self.kelas, 
                self.kegiatan, 
                self.parent.cbo_order_by.currentText(), 
                nama
            )
        except Exception as e:
            logger.error(f"Failed to get grade summary: {str(e)}")
            raise
        
    def load_default_setting(self):
        """Safely load default settings into UI controls"""
        try:
            # Block signals while updating
            for ctl in self.gabungan:
                ctl.blockSignals(True)
                
            # Apply default settings
            self.cbo_kertas.setCurrentText(self.DEFAULT_SETTING['kertas'])
            self.cbo_orientasi.setCurrentText(self.DEFAULT_SETTING['orientasi'])
            self.margin_left_spin.setValue(self.DEFAULT_SETTING['margin_left'])
            self.margin_top_spin.setValue(self.DEFAULT_SETTING['margin_top'])
            self.margin_right_spin.setValue(self.DEFAULT_SETTING['margin_right'])
            self.margin_bottom_spin.setValue(self.DEFAULT_SETTING['margin_bottom'])
            self.folder_rekap_plain.setPlainText(self.DEFAULT_SETTING['folder_rekap'])
            self.font_size_spin.setValue(self.DEFAULT_SETTING['font_size'])
            self.nilai_merah_spin.setValue(self.DEFAULT_SETTING['kkm'])
            self.kolom_nama_spin.setValue(self.DEFAULT_SETTING['kolom_nama'])
            self.kolom_pelajaran_spin.setValue(self.DEFAULT_SETTING['kolom_pelajaran'])
            self.ayah_radio.setChecked(self.DEFAULT_SETTING['ayah'])
            self.ibu_radio.setChecked(self.DEFAULT_SETTING['ibu'])
            self.alamat_radio.setChecked(self.DEFAULT_SETTING['alamat'])
            self.ayah_spin.setValue(self.DEFAULT_SETTING['kolom_ayah'])
            self.ibu_spin.setValue(self.DEFAULT_SETTING['kolom_ibu'])
            self.alamat_spin.setValue(self.DEFAULT_SETTING['kolom_alamat'])
            self.tinggi_baris_spin.setValue(self.DEFAULT_SETTING['tinggi_baris'])
            self.tinggi_baris_spin_1.setValue(self.DEFAULT_SETTING['tinggi_baris_1'])
            self.tinggi_baris_spin_3.setValue(self.DEFAULT_SETTING['tinggi_baris_3'])
            self.tinggi_baris_spin_10.setValue(self.DEFAULT_SETTING['tinggi_baris_10'])
            
            # Store current settings
            self.current_settings = self.DEFAULT_SETTING.copy()
            
        except Exception as e:
            logger.error(f"Failed to load default settings: {str(e)}")
            raise
        finally:
            # Always restore signals
            for ctl in self.gabungan:
                ctl.blockSignals(False)

    def load_setting_rekap(self, id_kelas: int):
        """Load settings for specified class ID with robust error handling"""
        try:
            if not id_kelas:
                logger.info("No class ID provided, loading defaults")
                self.load_default_setting()
                return
                
            # Get settings from database
            setting_data = self.SQL.get_setting_rekap(id_kelas)
            
            if not setting_data:
                logger.warning(f"No settings data returned for class {id_kelas}")
                self.load_default_setting()
                return
                
            # Handle case where setting_data might be a string or dict
            if isinstance(setting_data, str):
                try:
                    setting_data = json.loads(setting_data)
                except json.JSONDecodeError:
                    logger.error(f"Invalid JSON in settings for class {id_kelas}")
                    self.load_default_setting()
                    return
                    
            # Extract the actual settings
            if isinstance(setting_data, dict):
                settings = setting_data.get('setting_rekap_nilai')
            else:
                logger.error(f"Unexpected settings format for class {id_kelas}")
                self.load_default_setting()
                return
                
            # Parse settings if they're in JSON string format
            if isinstance(settings, str):
                try:
                    settings = json.loads(settings)
                except json.JSONDecodeError:
                    logger.error(f"Invalid nested JSON in settings for class {id_kelas}")
                    self.load_default_setting()
                    return
                    
            # Validate the settings structure
            if not isinstance(settings, dict) or not self._validate_settings(settings):
                logger.warning(f"Invalid settings structure for class {id_kelas}")
                self.load_default_setting()
                return
                
            # Apply the validated settings
            self._apply_settings_to_ui(settings)
            self.current_settings = settings
            logger.info(f"Successfully loaded settings for class {id_kelas}")
            
        except Exception as e:
            logger.error(f"Failed to load settings for class {id_kelas}: {str(e)}")
            self.load_default_setting()


    def _apply_settings_to_ui(self, settings: Dict):
        """Apply validated settings to UI controls"""
        try:
            # Block signals while updating
            for ctl in self.gabungan:
                ctl.blockSignals(True)
                
            # Apply settings with fallback to defaults
            self.cbo_kertas.setCurrentText(settings.get('kertas', self.DEFAULT_SETTING['kertas']))
            self.cbo_orientasi.setCurrentText(settings.get('orientasi', self.DEFAULT_SETTING['orientasi']))
            self.margin_left_spin.setValue(float(settings.get('margin_left', self.DEFAULT_SETTING['margin_left'])))
            self.margin_top_spin.setValue(float(settings.get('margin_top', self.DEFAULT_SETTING['margin_top'])))
            self.margin_right_spin.setValue(float(settings.get('margin_right', self.DEFAULT_SETTING['margin_right'])))
            self.margin_bottom_spin.setValue(float(settings.get('margin_bottom', self.DEFAULT_SETTING['margin_bottom'])))
            self.folder_rekap_plain.setPlainText(settings.get('folder_rekap', self.DEFAULT_SETTING['folder_rekap']))
            self.font_size_spin.setValue(float(settings.get('font_size', self.DEFAULT_SETTING['font_size'])))
            self.nilai_merah_spin.setValue(float(settings.get('kkm', self.DEFAULT_SETTING['kkm'])))
            self.kolom_nama_spin.setValue(float(settings.get('kolom_nama', self.DEFAULT_SETTING['kolom_nama'])))
            self.kolom_pelajaran_spin.setValue(float(settings.get('kolom_pelajaran', self.DEFAULT_SETTING['kolom_pelajaran'])))
            self.ayah_radio.setChecked(bool(settings.get('ayah', self.DEFAULT_SETTING['ayah'])))
            self.ibu_radio.setChecked(bool(settings.get('ibu', self.DEFAULT_SETTING['ibu'])))
            self.alamat_radio.setChecked(bool(settings.get('alamat', self.DEFAULT_SETTING['alamat'])))
            self.ayah_spin.setValue(float(settings.get('kolom_ayah', self.DEFAULT_SETTING['kolom_ayah'])))
            self.ibu_spin.setValue(float(settings.get('kolom_ibu', self.DEFAULT_SETTING['kolom_ibu'])))
            self.alamat_spin.setValue(float(settings.get('kolom_alamat', self.DEFAULT_SETTING['kolom_alamat'])))
            self.tinggi_baris_spin.setValue(float(settings.get('tinggi_baris', self.DEFAULT_SETTING['tinggi_baris'])))
            self.tinggi_baris_spin_1.setValue(float(settings.get('tinggi_baris_1', self.DEFAULT_SETTING['tinggi_baris_1'])))
            self.tinggi_baris_spin_3.setValue(float(settings.get('tinggi_baris_3', self.DEFAULT_SETTING['tinggi_baris_3'])))
            self.tinggi_baris_spin_10.setValue(float(settings.get('tinggi_baris_10', self.DEFAULT_SETTING['tinggi_baris_10'])))
            
        except Exception as e:
            logger.error(f"Error applying settings to UI: {str(e)}")
            raise
        finally:
            # Always restore signals
            for ctl in self.gabungan:
                ctl.blockSignals(False)


    def _validate_settings(self, settings: Dict) -> bool:
        """Validate settings structure thoroughly"""
        if not isinstance(settings, dict):
            return False
            
        # Check all required keys exist and have correct types
        required_keys = {
            'kertas': str,
            'orientasi': str,
            'margin_left': (int, float),
            'margin_top': (int, float),
            'margin_right': (int, float),
            'margin_bottom': (int, float),
            'folder_rekap': str,
            'font_size': (int, float),
            'kkm': (int, float),
            'kolom_nama': (int, float),
            'kolom_pelajaran': (int, float),
            'ayah': bool,
            'ibu': bool,
            'alamat': bool,
            'kolom_ayah': (int, float),
            'kolom_ibu': (int, float),
            'kolom_alamat': (int, float),
            'tinggi_baris': (int, float),
            'tinggi_baris_1': (int, float),
            'tinggi_baris_3': (int, float),
            'tinggi_baris_10': (int, float)
        }
        
        for key, expected_type in required_keys.items():
            if key not in settings:
                return False
            if not isinstance(settings[key], expected_type):
                return False
                
        return True

    def setting_rekap(self) -> Dict:
        """Get current settings from UI"""
        return {
            'kertas': self.cbo_kertas.currentText(),
            'orientasi': self.cbo_orientasi.currentText(),
            'margin_left': self.margin_left_spin.value(),
            'margin_top': self.margin_top_spin.value(),
            'margin_right': self.margin_right_spin.value(),
            'margin_bottom': self.margin_bottom_spin.value(),
            'folder_rekap': self.folder_rekap_plain.toPlainText(),
            'font_size': self.font_size_spin.value(),
            'kkm': self.nilai_merah_spin.value(),
            'kolom_nama': self.kolom_nama_spin.value(),
            'kolom_pelajaran': self.kolom_pelajaran_spin.value(),
            'ayah': self.ayah_radio.isChecked(),
            'ibu':self.ibu_radio.isChecked(),
            'alamat':self.alamat_radio.isChecked(),
            'kolom_ayah': self.ayah_spin.value(),
            'kolom_ibu': self.ibu_spin.value(),
            'kolom_alamat':self.alamat_spin.value(),
            'tinggi_baris': self.tinggi_baris_spin.value(),
            'tinggi_baris_1' : self.tinggi_baris_spin_1.value(),
            'tinggi_baris_3' : self.tinggi_baris_spin_3.value(),
            'tinggi_baris_10' : self.tinggi_baris_spin_10.value(),
        }

    def update_setting(self):
        """Update settings for current class"""
        try:
            if not self.id_kelas:
                logger.warning("No class selected, cannot update settings")
                return
                
            current_settings = json.dumps(self.setting_rekap())
            success = self.SQL.update_setting_rekap(self.id_kelas, current_settings)
            
            if success:
                logger.info("Settings updated successfully")
            else:
                logger.warning("Failed to update settings")
        except Exception as e:
            logger.error(f"Failed to update settings: {str(e)}")
            self._show_error("Error", f"Gagal menyimpan pengaturan: {str(e)}")

    def convert_mysql_to_reportlab(self, data: List[Dict]) -> List[List]:
        """Convert MySQL data to ReportLab compatible format"""
        if not data:
            return []
            
        try:
            header = list(data[0].keys())
            header_tbl = header_for_table(header)
            table_data = [header_tbl]
            
            for row in data:
                table_data.append([format_cell_data(row[key], zero="") for key in header])
                
            return table_data
        except Exception as e:
            logger.error(f"Failed to convert data: {str(e)}")
            raise

    def safe_float(self, val: Any) -> float:
        """Safely convert value to float"""
        try:
            return float(val) if val not in [None, ''] else 0.0
        except (ValueError, TypeError):
            return 0.0

    def save_pdf(self):
        """Save generated PDF to file"""
        try:
            if not self.pdf_data:
                self._show_error("Error", "Tidak ada PDF yang tersedia untuk disimpan")
                return
                
            kelas = getattr(self, 'kelas', '')
            kegiatan = getattr(self, 'kegiatan', '')
            tapel = getattr(self, 'tapel', '')
            jenjang = getattr(self, 'jenjang', '')
            
            namafile = f'Rekap Nilai {kegiatan} Kelas {kelas} {jenjang} Tapel {tapel}'
            open_after_save = self.radio_auto_open.isChecked()
            
            save_pdf(self, self.pdf_data, namafile, open_after_save)
        except Exception as e:
            self._show_error("Error", f"Gagal menyimpan PDF: {str(e)}")
            logger.error(f"Failed to save PDF: {str(e)}")

    def print_pdf(self):
        """Print generated PDF"""
        try:
            if not self.pdf_data:
                self._show_error("Error", "Tidak ada PDF yang tersedia untuk dicetak")
                return
                
            print_with_foxit(self.pdf_data)
        except Exception as e:
            self._show_error("Error", f"Gagal mencetak PDF: {str(e)}")
            logger.error(f"Failed to print PDF: {str(e)}")

    def _show_error(self, title: str, message: str):
        """Show error message dialog"""
        QMessageBox.critical(self, title, message)

    def _handle_pdf_ready(self, pdf_data: bytes):
        """Handle PDF ready signal"""
        self.pdf_data = pdf_data
        self.viewer.loadPDF(pdf_data)
